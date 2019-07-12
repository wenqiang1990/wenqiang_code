# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re, os, xlrd, xlwt
from robot.api import logger
from xlrd import open_workbook
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from  openpyxl.workbook  import  Workbook   
from  openpyxl.writer.excel  import  ExcelWriter    
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime 

class V2(unittest.TestCase):
    def setUp(self):
        #self.driver = webdriver.Ie()
        #self.driver = webdriver.Chrome()
        self.driver = webdriver.Firefox()
        self.driver.implicitly_wait(30)
        self.base_url = "http://login.wltest.com/"
        self.driver.maximize_window()
        self.verificationErrors = []
        self.accept_next_alert = True 
  
    def test_v2(self):    
        driver = self.driver
        driver.get(self.base_url + "/?source=")
        driver.find_element_by_id("username").clear()
        driver.find_element_by_id("username").send_keys("0000034927")
        driver.find_element_by_id("loginPasswd").clear()
        driver.find_element_by_id("loginPasswd").send_keys("i#3ml%Kv$2M@V*r")
        #driver.find_element_by_id("VerificationCode").clear()
        #driver.find_element_by_id("VerificationCode").send_keys("hhhh")
        driver.find_element_by_xpath("/html/body/div/div[2]/div[1]/div[2]/form/button").click()
        wb=load_workbook('C:\\Users\\wenqiang.RFDOA\\Desktop\\zhuoyue.xlsx')  
        sheetnames = wb.get_sheet_names()                   
        ws = wb.get_sheet_by_name(sheetnames[0])
        ft=Font(name='Neo Sans Intel',size=11)
        timecode = datetime.datetime.now().strftime("%Y%m%d%H%M%S")  
        formcode=range(int(timecode),int(int(timecode) + 20 ))
        print formcode
        for i in range(len(formcode)):
            if i==0:
                continue
            ws.cell(row=i+1,column=2).value=formcode[i] #此处修改亚马逊北京第4、40列；测试卓越网为2、39列
            ws.cell(row=i+1,column=2).font=ft
            ws.cell(row=i+1,column=39).value=formcode[i]
            ws.cell(row=i+1,column=39).font=ft
        wb.save('C:\\Users\\wenqiang.RFDOA\\Desktop\\zhuoyue.xlsx')
        formcode_files=[]
        formcode_files.append(ws.cell(row=2,column=2).value)
        txt=open("D:\\selenium_test\\formcode_excel.txt","w")
        txt.truncate()
        txt.write(str(formcode_files[0]))#write方法的实参需要为string类型
        txt.close()  
        
        driver.find_element_by_link_text(u"数据").click()
        driver.find_element_by_link_text(u"数据管理").click()
        driver.find_element_by_link_text(u"信息导入").click()
        frame=driver.find_element_by_xpath("/html/body/div/div[1]/div[3]/div/iframe[2]")
        driver.switch_to_frame(frame)#定位表单            time.sleep(3)
        driver.find_element_by_id("DDLMerchantID_chzn").click()
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[1]/tr/td[1]/div/div/div/input").send_keys(u"(测试)卓越网")
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[1]/tr/td[1]/div/div/div/input").send_keys(Keys.ARROW_DOWN)
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[1]/tr/td[1]/div/div/div/input").send_keys(Keys.ENTER)
        driver.find_element_by_id("txtOrderCount").send_keys("20")#导入订单数量
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[2]/tr/td[1]/div/a/span").click()
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[2]/tr/td[1]/div/div/div/input").send_keys(u"北京分拣")
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[2]/tr/td[1]/div/div/div/input").send_keys(Keys.ENTER)
        driver.find_element_by_id("txtFile").send_keys('C:\\Users\\wenqiang.RFDOA\\Desktop\\zhuoyue.xlsx')
        driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/form/table/tbody[2]/tr/td[2]/input[2]").click()
        time.sleep(6)
        driver.switch_to_alert().accept()#关闭弹出窗口
        driver.find_element_by_id("SearchBtn").click()
        time.sleep(80)
               

    
    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

if __name__ == "__main__":
    # 构造测试集
    suite = unittest.TestSuite()
    suite.addTest(V2("test_v2"))
    # 执行测试
    runner = unittest.TextTestRunner()
    runner.run(suite)