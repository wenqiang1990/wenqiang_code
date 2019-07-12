import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import re
import xlsxwriter
import xlrd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from csv_excel import csv_to_xlsx_pd 

filecsv='D:\\report\\issues.csv'
filexlsx='D:\\report\\issues.xlsx'
csv_to_xlsx_pd(filecsv,filexlsx) 
df=pd.read_excel('D:\\report\\issues.xlsx')
loandata=pd.DataFrame(pd.read_excel('D:\\report\\issues.xlsx'))
workbook = xlrd.open_workbook(r'D:\\report\\issues.xlsx')
# 获取所有sheet
#print(workbook.sheet_names()) # [u'sheet1', u'sheet2']
#获取sheet1
sheet1_name= workbook.sheet_names()[0]
#print(sheet1_name)
# 根据sheet索引或者名称获取sheet内容
sheet1 = workbook.sheet_by_name('data')
# sheet的名称，行数，列数
print(sheet1.name,sheet1.nrows,sheet1.ncols)
#统计一列中字段出现的次数
l1=[]
def pie_dict(col):
    l = df.groupby(col)[col].count()
    l=str(l)
    l=l.split('\n')
    l1=l[1:-1]
    m=[]
    for x in l1:
        x=x.split(' ')
        while '' in x:
            x.remove('')
        m.append(x)
    for x in m:
        if len(x)==3:
            x.pop(0)                    
    global n
    n=dict(m)#列表转字典 
#饼状图
def pie_picture(colname):
    pie_dict(colname)
    keys = list(n.keys())
    values = list(n.values())
    print(keys)
    print(values)
    #l=(n.get('codeReviewed',0),n.get('New',0),n.get('Processing',0),n.get('Pending',0),n.get('Resolved',0),n.get('Closed',0))
    #图表字体为华文细黑，字号为15
    plt.rc('font', family='STXihei', size=15)
    #设置饼图中每个数据分类的颜色
    #colors = ["#99CC01","#FFFF01","#0000FE","#FE0000","#A6A6A6","#D9E021"]
    colors = ["blue","red","coral","green","yellow","orange"]
    #设置饼图中每个数据分类的名称
    #name=['codeReviewed','New','Processing','Pending','Resolved','Closed']
    #创建饼图，设置分类标签，颜色和图表起始位置等
    #plt.pie(l,labels=name,colors=colors,explode=(0, 0, 0.15, 0, 0, 0),autopct='%1.1f%%')    
    plt.axes(aspect=1)  # set this , Figure is round, otherwise it is an ellipse
    plt.pie(values,labels=keys,colors=colors,autopct='%1.1f%%',pctdistance=0.8,labeldistance=1.1,shadow=False)
    #添加图表标题
    plt.title('本月测试任务'+ colname +'占比')
    #添加图例，并设置显示位置
    #plt.legend(['codeReviewed','New','Processing','Pending','Resolved','Closed'], loc='upper left')
    plt.legend(keys, loc='best', ncol = 2, frameon=True, prop={'size':10})
    #保存图片
    global filename1
    filename1='D:\\report\\' + colname + '.png'
    l1.append(filename1)
    plt.savefig(filename1)
    print(l1)
    plt.clf()#清空画布，不清空会造成第二张图片覆盖第一张
    #显示图表
    #plt.show()
#柱状图    
def bar_picture(colname1,colname2):
    pie_dict(colname1)
    keys = list(n.keys())
    values = list(n.values())
    print(keys)
    print(values)
    #按用户等级grade字段对贷款金额进行求和汇总
    loan_grade=loandata.groupby(colname1)[colname2].agg(sum)
    #图表字体为华文细黑，字号为15
    plt.rc('font', family='STXihei', size=15)
    #创建一个一维数组赋值给a
    #创建柱状图，数据源为按用户等级汇总的贷款金额，设置颜色，透明度和外边框颜色
    plt.bar(keys,loan_grade,color='#99CC01',alpha=0.8,align='center',edgecolor='white')
    #设置x轴标签
    plt.xlabel(colname2)
    #设置y周标签
    plt.ylabel(colname1)
    #设置图表标题
    plt.title('不同' + colname1 + '的' + colname2 + '分布')
    #设置图例的文字和在图表中的位置
    plt.legend([colname1], loc='upper right',prop={'size':10})
    #设置背景网格线的颜色，样式，尺寸和透明度
    plt.grid(color='#95a5a6',linestyle='--', linewidth=1,axis='y',alpha=0.4)
    #设置数据分类名称
    plt.xticks(keys,keys)
    #显示图表
    global filename2
    filename2='D:\\report\\' + colname1 + '.png'
    #plt.show()  
    l1.append(filename2)
    plt.savefig(filename2)
    print(l1)
    plt.clf()#清空画布，不清空会造成第二张图片覆盖第一张      
    
'''
book = xlsxwriter.Workbook(filexlsx)
sheet = book.add_worksheet()
sheet.insert_image('A143','D:\\report\\filename.png',{'x_offset':15,'y_offset':10})
book.close()

'''
def picture_excel():
    y=sheet1.nrows+1
    wb = load_workbook(filexlsx)#此方法不会读取到图片
    ws1=wb.active
    l2=['A','J','S']
    i=0
    for x in l1:
        img = Image(x)
        print(i)
        coord=l2[i]+str(y)
        i=i+1
        print(coord)
        ws1.add_image(img, coord)
    # Save the file
    wb.save(filexlsx)
pie_picture("状态")    
pie_picture("指派给") 
bar_picture("项目","耗时") 
picture_excel()









