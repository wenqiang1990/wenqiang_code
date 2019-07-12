# -*- coding: utf-8 -*-
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import re
import xlsxwriter
import xlrd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from matplotlib.ticker import MultipleLocator, FormatStrFormatter
from csv_excel import csv_to_xlsx_pd 
import operator

filecsv='D:\\report\\issues1.csv'
filexlsx='D:\\report\\issues1.xlsx'
csv_to_xlsx_pd(filecsv,filexlsx)
workbook1 = load_workbook(r'D:\\report\\issues1.xlsx')
df=pd.read_excel('D:\\report\\issues1.xlsx')
loandata=pd.DataFrame(pd.read_excel('D:\\report\\issues1.xlsx'))
workbook = xlrd.open_workbook(r'D:\\report\\issues1.xlsx')
# 获取所有sheet
#print(workbook.sheet_names()) # [u'sheet1', u'sheet2']
#获取sheet1
sheet1_name= workbook.sheet_names()[0]
#print(sheet1_name)
# 根据sheet索引或者名称获取sheet内容
sheet1 = workbook.sheet_by_name('data')
# sheet的名称，行数，列数
print(sheet1.name,sheet1.nrows,sheet1.ncols)
#统计一列中字段出现的次数
l1=[]
def pie_dict(col):
    l = df.groupby(col)[col].count()
    l=str(l)
    l=l.split('\n')
    l1=l[1:-1]
    m=[]
    for x in l1:
        x=x.split(' ')
        while '' in x:
            x.remove('')
        m.append(x)
    for x in m:
        if len(x)==3:
            x.pop(0)                    
    global n
    n1=dict(m)#列表转字典 
    for i in n1.keys():
        n1[i]=int(n1[i])
    print(n1)
    #n2=sorted(n1.items(), key=lambda d:d[1], reverse = True)
    n2=sorted(n1.items(),key=operator.itemgetter(1))#这里改为按照item的第二个字符排序，即value排序
    n=dict(n2)
    print(n)

# 获得所有sheet的名称
#print(workbook.get_sheet_names())
# 根据sheet名字获得sheet
#a_sheet = workbook.get_sheet_by_name('data')
# 获得sheet名
#print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = workbook1.active 
# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)

# 循环删除包含容信产品的行
#for column in sheet.columns:
def del_row(m):
    for cell in list(sheet.columns)[8]:
        #for i in l1:
        if cell.value== m:
            A = sheet['A' + str(cell.row)]
            print({cell.column},{cell.row},{cell.value})  # 返回的数字就是int型 
            print(A.value)
            # 读取后drop
            data = pd.read_excel(filexlsx, sheet_name="data")
            mydata = data.drop(A.value, axis=0)
            # 保存新的数据
            book = load_workbook(filexlsx)
            writer = pd.ExcelWriter(filexlsx,engine='openpyxl')
            writer.book = book
            # 清除原来的数据
            idx = book.sheetnames.index('data')
            book.remove(book.worksheets[idx])
            book.create_sheet('data', idx)
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            mydata.to_excel(writer, "data")
            writer.save()
    print('过滤产品类型为容信产品的数据')


#柱状图    
def bar_picture(colname1):
    pie_dict(colname1)
    keys = list(n.keys())
    values = list(n.values())
    print(keys)
    print(values)
    x=keys
    y=values
    #按用户等级grade字段对贷款金额进行求和汇总
    #loan_grade=loandata.groupby(colname1)[colname2].agg(sum)
    #图表字体为华文细黑，字号为15
    plt.rc('font', family='STXihei', size=6)
    
    '''  
    width =0.25
    plt.xticks(np.arange(len(x)),x)
    plt.bar(np.arange(len(x)),y,width,color='b')
    '''
    
    #创建柱状图，数据源为按用户等级汇总的贷款金额，设置颜色，透明度和外边框颜色
    plt.bar(keys,values,color='#99CC01',alpha=0.8,align='center',edgecolor='white')
    #设置x轴标签
    plt.xlabel(colname1)
    #设置y周标签
    plt.ylabel("数目")
    #设置图表标题
    plt.title('研发人员名下bug分布')
    #设置图例的文字和在图表中的位置
    #plt.legend([colname1], loc='upper right',prop={'size':8})
    #plt.axis([0, 1.1, 0, 20])  #代表横纵坐标的取值范围，前两个为x轴范围，后两个为y轴范围
    #设置背景网格线的颜色，样式，尺寸和透明度
    plt.grid(color='#95a5a6',linestyle='--', linewidth=1,axis='y',alpha=0.4)#axis='y'
    #设置数据分类名称
    #plt.xticks(keys)
    #显示图表
    global filename2
    filename2='D:\\report\\' + colname1 + '.png' 
    l1.append(filename2)
    plt.savefig(filename2)
    print(l1)
    plt.show()
    plt.clf()#清空画布，不清空会造成第二张图片覆盖第一张 
    
def picture_excel():
    y=sheet1.nrows+1
    wb = load_workbook(filexlsx)#此方法不会读取到图片
    ws1=wb.active
    l2=['A','J','S']
    i=0
    for x in l1:
        img = Image(x)
        print(i)
        coord=l2[i]+str(y)
        i=i+1
        print(coord)
        ws1.add_image(img, coord)
    # Save the file
    wb.save(filexlsx)    
    
    
'''    
del_row('Y 杨帅')
del_row('L 卢波文')
del_row('Q 秦永静')
del_row('Z 朱旭华')
del_row('L 李军静')
del_row('Q 秦帅')
'''    
bar_picture("指派给")
picture_excel()  
    
